## anonymize_actions.py
# pip install openpyxl

from openpyxl import load_workbook
import re

INPUT_FILE = "log.xlsx"
OUTPUT_FILE = "log_anonymized.xlsx"

ACTION_COLUMN = "action"


def blur_square_brackets(text):
    """
    Заменяет ВСЕ содержимое в квадратных скобках
    [что угодно] -> [СКРЫТО]
    """
    return re.sub(r"\[.*?\]", "[СКРЫТО]", text)


def blur_after_first_gt(text):
    """
    Скрыть всё после первого >
    """
    parts = text.split(">", 1)

    if len(parts) < 2:
        return text

    return parts[0] + "> [СКРЫТО]"


def blur_after_third_gt(text):
    """
    Скрыть всё после третьего >
    """

    positions = [m.start() for m in re.finditer(r">", text)]

    if len(positions) < 3:
        return text

    third_pos = positions[2]

    return text[:third_pos + 1] + " [СКРЫТО]"


def process_action(text):

    if text is None:
        return text

    text = str(text)

    # =========================================================
    # 1) portal-rshb
    # скрываем всё в квадратных скобках
    # =========================================================
    if "portal-rshb" in text:
        text = blur_square_brackets(text)

    # =========================================================
    # 2) TrueConf
    # скрыть всё после первого >
    # =========================================================
    elif "TrueConf" in text:
        text = blur_after_first_gt(text)

    # =========================================================
    # 3) Начало звонка
    # =========================================================
    elif "Начало звонка" in text:
        text = blur_after_first_gt(text)

    # =========================================================
    # 4) Завершение звонка
    # =========================================================
    elif "Завершение звонка" in text:
        text = blur_after_first_gt(text)

    # =========================================================
    # 5) Календарь
    # скрыть всё в квадратных скобках
    # =========================================================
    elif "Календарь" in text:
        text = blur_square_brackets(text)

    # =========================================================
    # 6) sgo-ap750
    # скрыть всё после третьего >
    # =========================================================
    elif "sgo-ap750" in text:
        text = blur_after_third_gt(text)

    return text


def main():

    wb = load_workbook(INPUT_FILE)
    ws = wb.active

    # поиск колонки action
    action_col = None

    for cell in ws[1]:
        if cell.value == ACTION_COLUMN:
            action_col = cell.column
            break

    if action_col is None:
        raise Exception("Колонка action не найдена")

    changed = 0

    for row in range(2, ws.max_row + 1):

        cell = ws.cell(row=row, column=action_col)

        old_value = cell.value

        new_value = process_action(old_value)

        if old_value != new_value:
            cell.value = new_value
            changed += 1

    wb.save(OUTPUT_FILE)

    print("===================================")
    print("Готово")
    print(f"Изменено строк: {changed}")
    print(f"Файл сохранен: {OUTPUT_FILE}")
    print("===================================")


if __name__ == "__main__":
    main()
